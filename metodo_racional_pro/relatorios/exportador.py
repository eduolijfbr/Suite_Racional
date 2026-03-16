# -*- coding: utf-8 -*-
"""
Exportador de Resultados em Múltiplos Formatos
"""

import json
import csv
import os
from datetime import datetime


class ExportadorResultados:
    """Exporta resultados em múltiplos formatos"""
    
    def __init__(self):
        pass
    
    def exportar(self, dados, caminho, formato=None):
        """
        Exporta resultados no formato especificado.
        
        Parâmetros:
            dados: dict com resultados
            caminho: Caminho do arquivo de saída
            formato: Formato de exportação (detectado pela extensão se não informado)
        """
        if formato is None:
            _, ext = os.path.splitext(caminho)
            formato = ext.lower().replace('.', '')
            
        formato = formato.lower()
        
        if formato in ['gpkg', 'geopackage']:
            self.exportar_geopackage(dados, caminho)
        elif formato in ['shp', 'shapefile']:
            self.exportar_shapefile(dados, caminho)
        elif formato in ['geojson', 'json']:
            self.exportar_geojson(dados, caminho)
        elif formato == 'csv':
            self.exportar_csv(dados, caminho)
        elif formato in ['xlsx', 'excel']:
            self.exportar_excel(dados, caminho)
        elif formato == 'kml':
            self.exportar_kml(dados, caminho)
        else:
            raise ValueError(f"Formato não suportado: {formato}")
    
    def exportar_csv(self, dados, caminho):
        """Exporta tabela de resultados para CSV"""
        dados_entrada = dados.get('dados_entrada', {})
        
        # Preparar dados para exportação
        linhas = [
            ['Parâmetro', 'Valor', 'Unidade'],
            ['Data do Cálculo', datetime.now().strftime('%Y-%m-%d %H:%M'), ''],
            ['Método', 'Racional', ''],
            ['', '', ''],
            ['DADOS DE ENTRADA', '', ''],
            ['Distância', dados_entrada.get('distancia', 0), 'm'],
            ['Desnível', dados_entrada.get('desnivel', 0), 'm'],
            ['Tempo de Concentração', dados_entrada.get('tempo', 0), 'min'],
            ['Área', dados_entrada.get('area', 0), 'km²'],
            ['Coef. Escoamento', dados_entrada.get('coef_escoamento', 0), '-'],
            ['Rugosidade', dados_entrada.get('rugosidade', 0), '-'],
            ['Declividade', dados_entrada.get('declividade', 0), '%'],
            ['Tempo de Retorno', dados_entrada.get('tempo_retorno', 25), 'anos'],
            ['Intensidade', dados_entrada.get('intensidade', 0), 'mm/h'],
            ['', '', ''],
            ['RESULTADOS', '', ''],
            ['Vazão', dados.get('vazao', 0), 'm³/s'],
            ['Diâmetro Círculo', dados.get('diametro', 0), 'm'],
            ['Lado Galeria Quadrada', dados.get('lado_galeria', 0), 'm'],
            ['Área da Seção', dados.get('area_secao', 0), 'm²'],
            ['Velocidade', dados.get('velocidade', 0), 'm/s'],
            ['Número de Froude', dados.get('froude', 0), '-'],
            ['Lâmina/Altura', dados.get('lamina', 0.85) * 100, '%'],
            ['', '', ''],
            ['STATUS', dados.get('status', ''), ''],
        ]
        
        with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerows(linhas)
    
    def exportar_excel(self, dados, caminho):
        """Exporta para Excel com formatação"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise ImportError(
                "Biblioteca openpyxl não instalada. "
                "Execute: pip install openpyxl"
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        dados_entrada = dados.get('dados_entrada', {})
        
        # Estilos
        titulo_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        # Título
        ws['A1'] = "RELATÓRIO DE CÁLCULO - MÉTODO RACIONAL"
        ws['A1'].font = titulo_font
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Dados de entrada
        ws['A4'] = "DADOS DE ENTRADA"
        ws['A4'].font = header_font
        
        row = 5
        ws.cell(row=row, column=1, value="Parâmetro").font = header_font
        ws.cell(row=row, column=2, value="Valor").font = header_font
        ws.cell(row=row, column=3, value="Unidade").font = header_font
        
        entradas = [
            ('Distância', dados_entrada.get('distancia', 0), 'm'),
            ('Desnível', dados_entrada.get('desnivel', 0), 'm'),
            ('Tempo de Concentração', dados_entrada.get('tempo', 0), 'min'),
            ('Área', dados_entrada.get('area', 0), 'km²'),
            ('Coef. Escoamento', dados_entrada.get('coef_escoamento', 0), '-'),
            ('Rugosidade', dados_entrada.get('rugosidade', 0), '-'),
            ('Declividade', dados_entrada.get('declividade', 0), '%'),
            ('Tempo de Retorno', dados_entrada.get('tempo_retorno', 25), 'anos'),
            ('Intensidade', dados_entrada.get('intensidade', 0), 'mm/h'),
        ]
        
        for i, (param, valor, unidade) in enumerate(entradas, start=6):
            ws.cell(row=i, column=1, value=param)
            ws.cell(row=i, column=2, value=valor)
            ws.cell(row=i, column=3, value=unidade)
        
        # Resultados
        row = len(entradas) + 8
        ws.cell(row=row, column=1, value="RESULTADOS").font = header_font
        
        row += 1
        ws.cell(row=row, column=1, value="Parâmetro").font = header_font
        ws.cell(row=row, column=2, value="Valor").font = header_font
        ws.cell(row=row, column=3, value="Unidade").font = header_font
        
        resultados = [
            ('Vazão', dados.get('vazao', 0), 'm³/s'),
            ('Diâmetro Círculo', dados.get('diametro', 0), 'm'),
            ('Lado Galeria Quadrada', dados.get('lado_galeria', 0), 'm'),
            ('Área da Seção', dados.get('area_secao', 0), 'm²'),
            ('Velocidade', dados.get('velocidade', 0), 'm/s'),
            ('Número de Froude', dados.get('froude', 0), '-'),
            ('Lâmina/Altura', dados.get('lamina', 0.85) * 100, '%'),
        ]
        
        for i, (param, valor, unidade) in enumerate(resultados, start=row+1):
            ws.cell(row=i, column=1, value=param)
            ws.cell(row=i, column=2, value=round(valor, 4))
            ws.cell(row=i, column=3, value=unidade)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        
        wb.save(caminho)
    
    def exportar_geojson(self, dados, caminho):
        """Exporta para GeoJSON"""
        dados_entrada = dados.get('dados_entrada', {})
        
        geojson = {
            "type": "FeatureCollection",
            "features": [
                {
                    "type": "Feature",
                    "properties": {
                        "tipo": "resultado_calculo",
                        "metodo": "Racional",
                        "data_calculo": datetime.now().isoformat(),
                        "vazao_m3s": dados.get('vazao', 0),
                        "diametro_m": dados.get('diametro', 0),
                        "lado_galeria_m": dados.get('lado_galeria', 0),
                        "area_secao_m2": dados.get('area_secao', 0),
                        "velocidade_ms": dados.get('velocidade', 0),
                        "froude": dados.get('froude', 0),
                        "area_km2": dados_entrada.get('area', 0),
                        "tempo_retorno": dados_entrada.get('tempo_retorno', 25),
                        "status": dados.get('status', '')
                    },
                    "geometry": None
                }
            ]
        }
        
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(geojson, f, indent=2, ensure_ascii=False)
    
    def exportar_geopackage(self, dados, caminho):
        """Exporta para GeoPackage"""
        try:
            from qgis.core import (
                QgsVectorLayer, QgsFeature, QgsField,
                QgsVectorFileWriter, QgsCoordinateReferenceSystem
            )
            from qgis.PyQt.QtCore import QVariant
        except ImportError:
            # Fallback para GeoJSON
            caminho_json = caminho.replace('.gpkg', '.geojson')
            self.exportar_geojson(dados, caminho_json)
            return
        
        # Criar camada em memória
        layer = QgsVectorLayer("Point?crs=EPSG:4326", "resultados", "memory")
        provider = layer.dataProvider()
        
        # Adicionar campos
        campos = [
            QgsField("vazao_m3s", QVariant.Double),
            QgsField("diam_m", QVariant.Double),
            QgsField("lado_galer", QVariant.Double),
            QgsField("area_seccao", QVariant.Double),
            QgsField("velocidade_ms", QVariant.Double),
            QgsField("froude", QVariant.Double),
            QgsField("status", QVariant.String),
        ]
        provider.addAttributes(campos)
        layer.updateFields()
        
        # Adicionar feature
        feature = QgsFeature()
        feature.setAttributes([
            dados.get('vazao', 0),
            dados.get('diametro', 0),
            dados.get('lado_galeria', 0),
            dados.get('area_secao', 0),
            dados.get('velocidade', 0),
            dados.get('froude', 0),
            dados.get('status', '')
        ])
        provider.addFeature(feature)
        
        # Salvar
        QgsVectorFileWriter.writeAsVectorFormat(
            layer, caminho, "UTF-8",
            QgsCoordinateReferenceSystem("EPSG:4326"),
            "GPKG"
        )
    
    def exportar_shapefile(self, dados, caminho):
        """Exporta para Shapefile"""
        # Similar ao GeoPackage, mas formato SHP
        try:
            from qgis.core import (
                QgsVectorLayer, QgsFeature, QgsField,
                QgsVectorFileWriter, QgsCoordinateReferenceSystem
            )
            from qgis.PyQt.QtCore import QVariant
            
            layer = QgsVectorLayer("Point?crs=EPSG:4326", "resultados", "memory")
            provider = layer.dataProvider()
            
            campos = [
                QgsField("vazao", QVariant.Double),
                QgsField("diametro", QVariant.Double),
                QgsField("lado_galer", QVariant.Double),
                QgsField("area_seca", QVariant.Double),
                QgsField("velocid", QVariant.Double),
                QgsField("froude", QVariant.Double),
            ]
            provider.addAttributes(campos)
            layer.updateFields()
            
            feature = QgsFeature()
            feature.setAttributes([
                dados.get('vazao', 0),
                dados.get('diametro', 0),
                dados.get('lado_galeria', 0),
                dados.get('area_secao', 0),
                dados.get('velocidade', 0),
                dados.get('froude', 0),
            ])
            provider.addFeature(feature)
            
            QgsVectorFileWriter.writeAsVectorFormat(
                layer, caminho, "UTF-8",
                QgsCoordinateReferenceSystem("EPSG:4326"),
                "ESRI Shapefile"
            )
            
        except ImportError:
            # Fallback para CSV
            caminho_csv = caminho.replace('.shp', '.csv')
            self.exportar_csv(dados, caminho_csv)
    
    def exportar_kml(self, dados, caminho):
        """Exporta para KML (Google Earth)"""
        dados_entrada = dados.get('dados_entrada', {})
        
        kml = f'''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>Resultado Método Racional</name>
    <description>
      Vazão: {dados.get('vazao', 0):.2f} m³/s
      Diâmetro Círculo: {dados.get('diametro', 0):.2f} m
      Lado Galeria: {dados.get('lado_galeria', 0):.2f} m
      Área da Seção: {dados.get('area_secao', 0):.2f} m²
      Velocidade: {dados.get('velocidade', 0):.2f} m/s
      Área: {dados_entrada.get('area', 0):.4f} km²
      TR: {dados_entrada.get('tempo_retorno', 25)} anos
    </description>
  </Document>
</kml>'''
        
        with open(caminho, 'w', encoding='utf-8') as f:
            f.write(kml)
